"""
App Streamlit per la ricerca inversa CUP -> CIG.

Dato uno o più CUP, l'app:
1. Interroga il dataset open data ANAC "cup" (mappatura CIG<->CUP) per
   trovare tutti i CIG collegati a ciascun CUP indicato.
2. Per ciascun CIG trovato, interroga l'API ANAC getSmartCig per
   recuperare i dettagli: stazione appaltante, CPV, date di pubblicazione,
   codice risposta.
3. Presenta i risultati in una tabella navigabile e scaricabile.

Avvio:
    pip install -r requirements.txt
    streamlit run app_cup_cig.py
"""

import io
import time
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Configurazione
# --------------------------------------------------------------------------

CKAN_API_BASE = "https://dati.anticorruzione.it/opendata/api/3/action/"
CIG_API_BASE = "https://api.anticorruzione.it/apicig/1.0.0/getSmartCig/"

# Dataset ANAC utilizzati, con URL diretto noto e stabile (file ZIP contenente il CSV)
# come ripiego se la risoluzione dinamica via API CKAN dovesse fallire.
DATASET_CONFIG = {
    "cup": {
        "url_statico": "https://dati.anticorruzione.it/opendata/download/dataset/cup/filesystem/cup_csv.zip",
        "titolo": "Mappatura CIG↔CUP",
    },
    "quadro-economico": {
        "url_statico": "https://dati.anticorruzione.it/opendata/download/dataset/quadro-economico/filesystem/quadro-economico_csv.zip",
        "titolo": "Quadro economico",
    },
    "stati-avanzamento": {
        "url_statico": "https://dati.anticorruzione.it/opendata/download/dataset/stati-avanzamento/filesystem/stati-avanzamento_csv.zip",
        "titolo": "Stati di avanzamento lavori (SAL)",
    },
}

RICHIESTA_TIMEOUT = 20
DOWNLOAD_TIMEOUT = 180
PAUSA_TRA_RICHIESTE = 0.0
MAX_TENTATIVI = 3
MAX_WORKER_THREADS = 10  # chiamate all'API getSmartCig in parallelo

# Il portale open data ANAC filtra le richieste prive di intestazioni "da browser":
# senza uno User-Agent realistico alcune richieste vengono rifiutate (pagina di errore
# invece di JSON/CSV). Le stesse intestazioni vengono usate anche verso l'API getSmartCig
# per uniformità, senza effetti negativi.
HEADERS_BROWSER = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/csv, application/zip, */*",
}

COLONNE_RISULTATO = [
    "CUP",
    "CIG",
    "Codice risposta",
    "Stazione Appaltante",
    "Citta",
    "Regione",
    "Codice AUSA",
    "Data pubblicazione bando",
    "Data creazione CIG",
    "CPV",
]

COLORE_PRIMARIO = "#0B3D62"
COLORE_SECONDARIO = "#EEF3F8"

st.set_page_config(
    page_title="Ricerca CIG collegati a un CUP — ANAC",
    page_icon="🔎",
    layout="wide",
)

# --------------------------------------------------------------------------
# Stile (stesso stile dell'app CIG -> CUP)
# --------------------------------------------------------------------------

st.markdown(
    f"""
    <style>
    .main {{
        background-color: #FAFBFC;
    }}
    .app-header {{
        background-color: {COLORE_PRIMARIO};
        padding: 1.6rem 2rem;
        border-radius: 8px;
        margin-bottom: 1.5rem;
    }}
    .app-header h1 {{
        color: white;
        font-size: 1.5rem;
        margin: 0;
        font-weight: 600;
    }}
    .app-header p {{
        color: #D6E2EE;
        margin: 0.3rem 0 0 0;
        font-size: 0.95rem;
    }}
    .info-box {{
        background-color: {COLORE_SECONDARIO};
        border-left: 4px solid {COLORE_PRIMARIO};
        padding: 0.9rem 1.2rem;
        border-radius: 4px;
        margin-bottom: 1rem;
        font-size: 0.9rem;
    }}
    div[data-testid="stMetricValue"] {{
        font-size: 1.6rem;
        color: {COLORE_PRIMARIO};
    }}
    .stButton > button {{
        background-color: {COLORE_PRIMARIO};
        color: white;
        border-radius: 6px;
        padding: 0.5rem 1.5rem;
        border: none;
        font-weight: 600;
    }}
    .stButton > button:hover {{
        background-color: #0A3352;
        color: white;
    }}
    footer {{visibility: hidden;}}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="app-header">
        <h1>Ricerca CIG collegati a un CUP</h1>
        <p>Trova tutti i CIG associati a uno o più CUP e recupera i relativi dettagli dall'API ANAC</p>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="info-box">
    Questa app cerca i CIG collegati a un CUP incrociando il dataset open data ANAC
    "cup" (mappatura CIG↔CUP degli appalti ordinari), arricchisce ogni CIG trovato
    con i dettagli restituiti dall'API pubblica <b>getSmartCig</b> (stazione appaltante,
    categoria merceologica CPV, date di pubblicazione), e recupera inoltre le informazioni
    disponibili nei dataset ANAC <b>quadro economico</b> e <b>stati di avanzamento lavori
    (SAL)</b> per ciascun CIG trovato.
    </div>
    """,
    unsafe_allow_html=True,
)


# --------------------------------------------------------------------------
# Funzioni: recupero dataset CIG<->CUP
# --------------------------------------------------------------------------


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def risolvi_url_dataset(nome_dataset: str) -> str:
    """Restituisce l'URL di download del file di un dataset ANAC (cup,
    quadro-economico, stati-avanzamento, ...). Prova prima a chiedere
    l'URL aggiornato tramite l'API CKAN; se il portale rifiuta la
    richiesta (capita con alcune reti cloud), usa l'URL statico noto
    come ripiego."""
    try:
        resp = requests.get(
            CKAN_API_BASE + "package_show",
            params={"id": nome_dataset},
            headers=HEADERS_BROWSER,
            timeout=RICHIESTA_TIMEOUT,
        )
        resp.raise_for_status()
        dati = resp.json()
        risorse = dati.get("result", {}).get("resources", [])
        for r in risorse:
            nome = r.get("name", "").lower()
            formato = r.get("format", "").upper()
            if formato in ("CSV", "ZIP") and "log" not in nome and not nome[:8].isdigit():
                return r["url"]
        for r in risorse:
            if r.get("format", "").upper() in ("CSV", "ZIP"):
                return r["url"]
    except Exception:
        pass  # ripiego sull'URL statico

    return DATASET_CONFIG[nome_dataset]["url_statico"]


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def scarica_dataset_anac(nome_dataset: str) -> pd.DataFrame:
    """Scarica e carica in memoria un dataset ANAC (pensata per dataset
    contenuti come 'cup'). Scarica in streaming su disco per non raddoppiare
    l'uso di RAM, poi lo carica in un unico DataFrame."""
    import zipfile
    import tempfile
    import os

    url = risolvi_url_dataset(nome_dataset)
    resp = requests.get(url, headers=HEADERS_BROWSER, timeout=DOWNLOAD_TIMEOUT, stream=True)
    resp.raise_for_status()

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".bin")
    try:
        with os.fdopen(tmp_fd, "wb") as tmp_file:
            for blocco in resp.iter_content(chunk_size=1024 * 1024):
                if blocco:
                    tmp_file.write(blocco)

        if os.path.getsize(tmp_path) == 0:
            raise RuntimeError("Il portale ANAC ha restituito una risposta vuota.")

        if zipfile.is_zipfile(tmp_path):
            with zipfile.ZipFile(tmp_path) as zf:
                nomi_csv = [n for n in zf.namelist() if n.lower().endswith(".csv")]
                if not nomi_csv:
                    raise RuntimeError(f"Nessun file CSV trovato nello ZIP scaricato. Contenuto: {zf.namelist()}")
                with zf.open(nomi_csv[0]) as f_csv:
                    prima_riga = f_csv.readline().decode("utf-8", errors="replace")
                with zf.open(nomi_csv[0]) as f_csv:
                    separatore = _rileva_separatore(prima_riga)
                    df = pd.read_csv(f_csv, dtype=str, sep=separatore, low_memory=False)
        else:
            with open(tmp_path, "rb") as f_peek:
                testa = f_peek.read(200).lstrip()
            if testa[:1] in (b"<", b"{"):
                testo_anteprima = testa.decode("utf-8", errors="replace")
                raise RuntimeError(
                    f"Il portale ANAC ha restituito una pagina di errore invece del file. Anteprima: {testo_anteprima}"
                )
            with open(tmp_path, "r", encoding="utf-8", errors="replace") as f_peek2:
                prima_riga = f_peek2.readline()
            separatore = _rileva_separatore(prima_riga)
            df = pd.read_csv(tmp_path, dtype=str, sep=separatore, low_memory=False)

        df.columns = [c.strip().upper() for c in df.columns]
        return df
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def leggi_dataset_da_file_caricato(contenuto: bytes) -> pd.DataFrame:
    """Legge un dataset ANAC (CSV o ZIP con CSV) da un file caricato manualmente."""
    import zipfile

    if zipfile.is_zipfile(io.BytesIO(contenuto)):
        with zipfile.ZipFile(io.BytesIO(contenuto)) as zf:
            nomi_csv = [n for n in zf.namelist() if n.lower().endswith(".csv")]
            with zf.open(nomi_csv[0]) as f_csv:
                df = pd.read_csv(f_csv, dtype=str, sep=None, engine="python")
    else:
        df = pd.read_csv(io.BytesIO(contenuto), dtype=str, sep=None, engine="python")
    df.columns = [c.strip().upper() for c in df.columns]
    return df


def _rileva_separatore(prima_riga: str) -> str:
    """Individua il separatore CSV (ANAC usa spesso ';') senza affidarsi
    al motore 'python' di pandas su file di grandi dimensioni, che è lento."""
    import csv as _csv
    try:
        return _csv.Sniffer().sniff(prima_riga, delimiters=";,\t|").delimiter
    except Exception:
        return ";" if prima_riga.count(";") >= prima_riga.count(",") else ","


def scarica_e_filtra_dataset_a_blocchi(
    nome_dataset: str, valori_da_cercare: list, pattern_colonna: str
) -> pd.DataFrame:
    """Scarica un dataset ANAC (potenzialmente molto grande) in streaming su
    disco e lo filtra leggendolo a blocchi (chunk), senza mai caricarlo
    interamente in memoria. Evita i crash per esaurimento RAM che si
    verificano con i dataset più pesanti (es. quadro-economico,
    stati-avanzamento)."""
    import zipfile
    import tempfile
    import os

    url = risolvi_url_dataset(nome_dataset)
    resp = requests.get(url, headers=HEADERS_BROWSER, timeout=DOWNLOAD_TIMEOUT, stream=True)
    resp.raise_for_status()

    valori_set = set(str(v).strip().upper() for v in valori_da_cercare)

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".bin")
    try:
        with os.fdopen(tmp_fd, "wb") as tmp_file:
            for blocco in resp.iter_content(chunk_size=1024 * 1024):
                if blocco:
                    tmp_file.write(blocco)

        if os.path.getsize(tmp_path) == 0:
            raise RuntimeError("Il portale ANAC ha restituito una risposta vuota.")

        def _apri_csv_sorgente():
            """Restituisce (file_object_testuale, prima_riga) del CSV, gestendo lo ZIP."""
            if zipfile.is_zipfile(tmp_path):
                zf = zipfile.ZipFile(tmp_path)
                nomi_csv = [n for n in zf.namelist() if n.lower().endswith(".csv")]
                if not nomi_csv:
                    raise RuntimeError(f"Nessun file CSV nello ZIP di '{nome_dataset}'.")
                f = zf.open(nomi_csv[0])
                prima_riga_bytes = f.readline()
                f.close()
                f = zf.open(nomi_csv[0])
                return f, prima_riga_bytes.decode("utf-8", errors="replace")
            else:
                with open(tmp_path, "rb") as f_peek:
                    testa = f_peek.read(200).lstrip()
                if testa[:1] in (b"<", b"{"):
                    testo_anteprima = testa.decode("utf-8", errors="replace")
                    raise RuntimeError(
                        f"Il portale ANAC ha restituito una pagina di errore invece del file. Anteprima: {testo_anteprima}"
                    )
                f = open(tmp_path, "rb")
                prima_riga_bytes = f.readline()
                f.seek(0)
                return f, prima_riga_bytes.decode("utf-8", errors="replace")

        f_sorgente, prima_riga = _apri_csv_sorgente()
        separatore = _rileva_separatore(prima_riga)

        pezzi_trovati = []
        colonna_chiave = None
        try:
            for chunk_df in pd.read_csv(
                f_sorgente, dtype=str, sep=separatore, chunksize=50_000, low_memory=False
            ):
                chunk_df.columns = [c.strip().upper() for c in chunk_df.columns]
                if colonna_chiave is None:
                    colonna_chiave = next(
                        (c for c in chunk_df.columns if c == pattern_colonna or pattern_colonna in c),
                        None,
                    )
                    if colonna_chiave is None:
                        raise RuntimeError(
                            f"Colonna '{pattern_colonna}' non trovata nel dataset '{nome_dataset}'. "
                            f"Colonne disponibili: {list(chunk_df.columns)}"
                        )
                match = chunk_df[chunk_df[colonna_chiave].astype(str).str.strip().str.upper().isin(valori_set)]
                if not match.empty:
                    pezzi_trovati.append(match.copy())
        finally:
            f_sorgente.close()

        if pezzi_trovati:
            return pd.concat(pezzi_trovati, ignore_index=True)
        return pd.DataFrame()

    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def trova_cig_per_cup(df_cup: pd.DataFrame, cup_list: list) -> pd.DataFrame:
    """Filtra il dataset CIG<->CUP per i CUP richiesti."""
    colonna_cup = next((c for c in df_cup.columns if "CUP" in c), None)
    colonna_cig = next((c for c in df_cup.columns if "CIG" in c), None)
    if colonna_cup is None or colonna_cig is None:
        raise RuntimeError(
            f"Colonne CUP/CIG non riconosciute nel dataset. Colonne trovate: {list(df_cup.columns)}"
        )
    cup_set = set(c.strip().upper() for c in cup_list)
    df_filtrato = df_cup[df_cup[colonna_cup].str.strip().str.upper().isin(cup_set)]
    return df_filtrato[[colonna_cup, colonna_cig]].rename(
        columns={colonna_cup: "CUP", colonna_cig: "CIG"}
    )


def filtra_per_cig(df: pd.DataFrame, cig_list: list) -> pd.DataFrame:
    """Filtra un dataset ANAC generico per una lista di CIG, individuando
    automaticamente la colonna che contiene il CIG."""
    colonna_cig = next((c for c in df.columns if c == "CIG" or c.endswith("_CIG") or c.startswith("CIG_")), None)
    if colonna_cig is None:
        colonna_cig = next((c for c in df.columns if "CIG" in c), None)
    if colonna_cig is None:
        raise RuntimeError(f"Colonna CIG non riconosciuta nel dataset. Colonne trovate: {list(df.columns)}")
    cig_set = set(c.strip().upper() for c in cig_list)
    return df[df[colonna_cig].str.strip().str.upper().isin(cig_set)]


# --------------------------------------------------------------------------
# Funzioni: arricchimento dettagli CIG via getSmartCig
# --------------------------------------------------------------------------


def interroga_cig(cig: str) -> dict:
    import random

    url = CIG_API_BASE + cig
    ultimo_errore = None
    for tentativo in range(1, MAX_TENTATIVI + 1):
        try:
            resp = requests.get(url, headers=HEADERS_BROWSER, timeout=RICHIESTA_TIMEOUT)
            if resp.status_code >= 500:
                ultimo_errore = f"{resp.status_code} Server Error"
                time.sleep(min(2 ** tentativo, 6) + random.uniform(0, 0.5))
                continue
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            ultimo_errore = e
            time.sleep(min(2 ** tentativo, 6) + random.uniform(0, 0.5))
    return {"errore": str(ultimo_errore)}


def estrai_dati(cup: str, cig: str, risposta: dict) -> dict:
    riga = {
        "CUP": cup,
        "CIG": cig,
        "Codice risposta": risposta.get("codice_risposta", ""),
        "Stazione Appaltante": "",
        "Citta": "",
        "Regione": "",
        "Codice AUSA": "",
        "Data pubblicazione bando": "",
        "Data creazione CIG": "",
        "CPV": "",
    }

    if "errore" in risposta:
        riga["Codice risposta"] = f"ERRORE RETE: {risposta['errore']}"
        return riga

    bando = risposta.get("bando") or {}
    cpv_list = bando.get("CPV") or []
    cpv_desc = []
    for c in cpv_list:
        cod = c.get("COD_CPV", "")
        desc = c.get("DESCRIZIONE_CPV", "")
        if cod or desc:
            cpv_desc.append(f"{cod} - {desc}")
    riga["CPV"] = "; ".join(cpv_desc)

    sa = risposta.get("stazione_appaltante") or {}
    riga["Stazione Appaltante"] = sa.get("DENOMINAZIONE_AMMINISTRAZIONE_APPALTANTE", "")
    riga["Citta"] = sa.get("CITTA", "")
    riga["Regione"] = sa.get("REGIONE", "")
    riga["Codice AUSA"] = sa.get("CODICE_AUSA", "")

    pub = risposta.get("pubblicazioni") or {}
    riga["Data pubblicazione bando"] = pub.get("DATA_PUBBLICAZIONE", "")
    riga["Data creazione CIG"] = pub.get("DATA_CREAZIONE", "")

    return riga


def _scrivi_foglio_dinamico(ws, df: pd.DataFrame):
    """Scrive un DataFrame con colonne dinamiche (non note a priori) in un foglio Excel."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0B3D62", end_color="0B3D62", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for i in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.freeze_panes = "A2"


def genera_excel_risultato(df: pd.DataFrame, fogli_extra=None) -> bytes:
    """Genera il file Excel finale. 'fogli_extra' è un dizionario opzionale
    {nome_foglio: dataframe} per aggiungere fogli con colonne dinamiche
    (es. quadro economico, stati di avanzamento)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CIG collegati a CUP"

    for col_idx, col_name in enumerate(COLONNE_RISULTATO, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0B3D62", end_color="0B3D62", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    larghezze = [16, 16, 14, 42, 20, 18, 12, 20, 20, 45]
    for i, larg in enumerate(larghezze, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    ws.freeze_panes = "A2"

    if fogli_extra:
        for nome_foglio, df_extra in fogli_extra.items():
            if df_extra is None or df_extra.empty:
                continue
            ws_extra = wb.create_sheet(nome_foglio[:31])  # Excel limita a 31 caratteri
            _scrivi_foglio_dinamico(ws_extra, df_extra)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def ottieni_dataset_con_fallback(nome_dataset: str, obbligatorio: bool = True):
    """Gestisce il download di un dataset ANAC con fallback al caricamento
    manuale in caso di errore. Usa st.session_state per ricordare il
    risultato tra un rerun e l'altro. Restituisce il DataFrame oppure None."""
    chiave_df = f"df_{nome_dataset}"
    chiave_errore = f"errore_{nome_dataset}"
    titolo = DATASET_CONFIG[nome_dataset]["titolo"]

    if chiave_df not in st.session_state:
        try:
            st.session_state[chiave_df] = scarica_dataset_anac(nome_dataset)
            st.session_state[chiave_errore] = None
        except Exception as e:
            st.session_state[chiave_df] = None
            st.session_state[chiave_errore] = str(e)

    if st.session_state.get(chiave_errore):
        livello = st.error if obbligatorio else st.warning
        livello(
            f"Impossibile scaricare automaticamente il dataset ANAC '{titolo}': "
            f"{st.session_state[chiave_errore]}"
        )
        st.markdown(
            f"🔗 [Apri il dataset '{nome_dataset}' su dati.anticorruzione.it]"
            f"(https://dati.anticorruzione.it/opendata/dataset/{nome_dataset}) "
            "— scaricalo e ricaricalo qui sotto per proseguire."
        )
        file_manuale = st.file_uploader(
            f"Carica il file '{titolo}' (CSV o ZIP)",
            type=["csv", "zip"],
            key=f"upload_{nome_dataset}",
        )
        if file_manuale is not None:
            try:
                df_manuale = leggi_dataset_da_file_caricato(file_manuale.read())
                st.session_state[chiave_df] = df_manuale
                st.session_state[chiave_errore] = None
                st.success("File caricato correttamente.")
                st.rerun()
            except Exception as e2:
                st.error(f"Impossibile leggere il file caricato: {e2}")

    return st.session_state.get(chiave_df)


def ottieni_dataset_filtrato_con_fallback(
    nome_dataset: str, valori_da_cercare: list, pattern_colonna: str, obbligatorio: bool = False
):
    """Come ottieni_dataset_con_fallback, ma per i dataset di grandi
    dimensioni: scarica e filtra a blocchi (senza mai caricare tutto il
    file in memoria) e mette in cache solo il risultato già filtrato,
    molto più leggero."""
    chiave_df = f"df_filtrato_{nome_dataset}_{hash(tuple(sorted(valori_da_cercare)))}"
    chiave_errore = f"errore_{nome_dataset}"
    titolo = DATASET_CONFIG[nome_dataset]["titolo"]

    if chiave_df not in st.session_state:
        try:
            st.session_state[chiave_df] = scarica_e_filtra_dataset_a_blocchi(
                nome_dataset, valori_da_cercare, pattern_colonna
            )
            st.session_state[chiave_errore] = None
        except Exception as e:
            st.session_state[chiave_df] = None
            st.session_state[chiave_errore] = str(e)

    if st.session_state.get(chiave_errore):
        livello = st.error if obbligatorio else st.warning
        livello(
            f"Impossibile recuperare il dataset ANAC '{titolo}': "
            f"{st.session_state[chiave_errore]}"
        )
        st.markdown(
            f"🔗 [Apri il dataset '{nome_dataset}' su dati.anticorruzione.it]"
            f"(https://dati.anticorruzione.it/opendata/dataset/{nome_dataset}) "
            "— scaricalo e ricaricalo qui sotto per proseguire."
        )
        file_manuale = st.file_uploader(
            f"Carica il file '{titolo}' (CSV o ZIP)",
            type=["csv", "zip"],
            key=f"upload_filtrato_{nome_dataset}",
        )
        if file_manuale is not None:
            try:
                df_completo_manuale = leggi_dataset_da_file_caricato(file_manuale.read())
                st.session_state[chiave_df] = filtra_per_cig(df_completo_manuale, valori_da_cercare)
                st.session_state[chiave_errore] = None
                st.success("File caricato correttamente.")
                st.rerun()
            except Exception as e2:
                st.error(f"Impossibile leggere il file caricato: {e2}")

    return st.session_state.get(chiave_df)


def genera_template_excel_cup() -> bytes:
    """Genera un file Excel vuoto con la sola intestazione 'CUP', da usare
    come modello per il caricamento."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LISTA CUP"
    cell = ws.cell(row=1, column=1, value="CUP")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="0B3D62", end_color="0B3D62", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 24
    # righe di esempio, facoltative, cancellabili dall'utente
    ws.cell(row=2, column=1, value="F16H21000000008")
    ws.cell(row=3, column=1, value="F55G21000000001")
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------
# Interfaccia
# --------------------------------------------------------------------------

col_input, col_info = st.columns([2, 1])

with col_input:
    tab_testo, tab_file = st.tabs(["Incolla i CUP", "Carica file Excel"])

    cup_list = []

    with tab_testo:
        testo_cup = st.text_area(
            "Inserisci uno o più CUP (uno per riga)",
            height=140,
            placeholder="F16H21000000008\nF55G21000000001\n...",
        )
        if testo_cup.strip():
            cup_list = [c.strip() for c in testo_cup.splitlines() if c.strip()]

    with tab_file:
        col_uploader, col_template = st.columns([3, 1])
        with col_uploader:
            file_caricato = st.file_uploader("File Excel con una colonna CUP", type=["xlsx"])
        with col_template:
            st.write("")
            st.write("")
            st.download_button(
                label="Scarica template",
                data=genera_template_excel_cup(),
                file_name="template_lista_cup.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        if file_caricato is not None:
            try:
                wb_up = openpyxl.load_workbook(file_caricato, data_only=True)
                ws_up = wb_up[wb_up.sheetnames[0]]
                intestazioni = [c.value for c in next(ws_up.iter_rows(min_row=1, max_row=1))]
                colonna_idx = None
                for idx, val in enumerate(intestazioni, start=1):
                    if val and "CUP" in str(val).upper():
                        colonna_idx = idx
                        break
                if colonna_idx is None:
                    colonna_idx = 1
                for row in ws_up.iter_rows(min_row=2, values_only=True):
                    val = row[colonna_idx - 1]
                    if val:
                        cup_list.append(str(val).strip())
            except Exception as e:
                st.error(f"Impossibile leggere il file: {e}")

with col_info:
    st.markdown("**Informazioni**")
    st.caption("Fonte mappatura CIG↔CUP: dataset open data ANAC 'cup'")
    st.caption("Fonte dettagli CIG: API pubblica ANAC — getSmartCig")
    st.caption("Fonte quadro economico: dataset open data ANAC 'quadro-economico'")
    st.caption("Fonte SAL: dataset open data ANAC 'stati-avanzamento'")
    st.caption("I dataset ANAC vengono scaricati una volta e riutilizzati per 24 ore")

if cup_list:
    cup_list = list(dict.fromkeys(cup_list))  # rimuove duplicati mantenendo l'ordine
    st.success(f"**{len(cup_list)}** CUP pronti per la ricerca.")

    avvia = st.button("Cerca CIG collegati", type="primary")

    if avvia:
        # forza un nuovo tentativo di download ad ogni click sul pulsante
        for chiave in ("df_cup", "errore_cup"):
            st.session_state.pop(chiave, None)
        st.session_state["ricerca_avviata"] = True

    if st.session_state.get("ricerca_avviata"):
        with st.spinner("Recupero il dataset CIG↔CUP di ANAC (può richiedere qualche secondo)..."):
            df_cup_dataset = ottieni_dataset_con_fallback("cup", obbligatorio=True)

        if df_cup_dataset is not None:
            corrispondenze = trova_cig_per_cup(df_cup_dataset, cup_list)

            cup_trovati = set(corrispondenze["CUP"].str.upper())
            cup_non_trovati = [c for c in cup_list if c.upper() not in cup_trovati]

            if corrispondenze.empty:
                st.warning("Nessun CIG trovato per i CUP indicati nel dataset ANAC.")
            else:
                st.info(
                    f"Trovati **{len(corrispondenze)}** CIG collegati a "
                    f"**{len(cup_trovati)}** CUP su {len(cup_list)} richiesti."
                )

                barra_avanzamento = st.progress(0, text="Recupero dettagli CIG in corso...")
                placeholder_tabella = st.empty()

                righe = corrispondenze.to_dict("records")
                risultati = [None] * len(righe)
                completati = 0

                with ThreadPoolExecutor(max_workers=min(MAX_WORKER_THREADS, len(righe))) as executor:
                    future_to_idx = {
                        executor.submit(interroga_cig, riga_match["CIG"]): idx
                        for idx, riga_match in enumerate(righe)
                    }
                    for future in as_completed(future_to_idx):
                        idx = future_to_idx[future]
                        cup = righe[idx]["CUP"]
                        cig = righe[idx]["CIG"]
                        risposta = future.result()
                        risultati[idx] = estrai_dati(cup, cig, risposta)

                        completati += 1
                        barra_avanzamento.progress(
                            completati / len(righe),
                            text=f"CIG elaborati: {completati} di {len(righe)}",
                        )

                        df_parziale = pd.DataFrame(
                            [r for r in risultati if r is not None], columns=COLONNE_RISULTATO
                        )
                        placeholder_tabella.dataframe(df_parziale, use_container_width=True, height=350)

                barra_avanzamento.empty()
                df_finale = pd.DataFrame(risultati, columns=COLONNE_RISULTATO)
                lista_cig_trovati = df_finale["CIG"].dropna().unique().tolist()

                st.markdown("### Riepilogo per CUP")
                riepilogo = (
                    df_finale.groupby("CUP")["CIG"]
                    .nunique()
                    .reset_index()
                    .rename(columns={"CIG": "Numero CIG collegati"})
                )
                if cup_non_trovati:
                    df_non_trovati = pd.DataFrame(
                        {"CUP": cup_non_trovati, "Numero CIG collegati": 0}
                    )
                    riepilogo = pd.concat([riepilogo, df_non_trovati], ignore_index=True)
                st.dataframe(riepilogo, use_container_width=True, height=min(400, 40 + 35 * len(riepilogo)))

                st.markdown("### Dettaglio completo CIG")
                st.dataframe(df_finale, use_container_width=True, height=400)

                # ------------------------------------------------------------
                # Quadro economico
                # ------------------------------------------------------------
                st.markdown("### Quadro economico")
                with st.spinner("Recupero le voci di quadro economico per i CIG trovati..."):
                    df_qe_filtrato = ottieni_dataset_filtrato_con_fallback(
                        "quadro-economico", lista_cig_trovati, "CIG", obbligatorio=False
                    )
                if df_qe_filtrato is None:
                    df_qe_filtrato = pd.DataFrame()
                elif df_qe_filtrato.empty:
                    st.caption("Nessuna voce di quadro economico trovata per i CIG individuati.")
                else:
                    st.dataframe(df_qe_filtrato, use_container_width=True, height=300)

                # ------------------------------------------------------------
                # Stati di avanzamento lavori (SAL)
                # ------------------------------------------------------------
                st.markdown("### Stati di avanzamento lavori (SAL)")
                with st.spinner("Recupero gli stati di avanzamento per i CIG trovati..."):
                    df_sal_filtrato = ottieni_dataset_filtrato_con_fallback(
                        "stati-avanzamento", lista_cig_trovati, "CIG", obbligatorio=False
                    )
                if df_sal_filtrato is None:
                    df_sal_filtrato = pd.DataFrame()
                elif df_sal_filtrato.empty:
                    st.caption("Nessuno stato di avanzamento lavori trovato per i CIG individuati.")
                else:
                    st.dataframe(df_sal_filtrato, use_container_width=True, height=300)

                # ------------------------------------------------------------
                # Download
                # ------------------------------------------------------------
                fogli_extra = {
                    "Quadro economico": df_qe_filtrato,
                    "Stati avanzamento (SAL)": df_sal_filtrato,
                }
                excel_bytes = genera_excel_risultato(df_finale, fogli_extra=fogli_extra)
                csv_bytes = df_finale.to_csv(index=False).encode("utf-8-sig")

                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button(
                        label="Scarica risultati completi (Excel, più fogli)",
                        data=excel_bytes,
                        file_name="cig_collegati_a_cup.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                with col_dl2:
                    st.download_button(
                        label="Scarica solo dettaglio CIG (CSV)",
                        data=csv_bytes,
                        file_name="cig_collegati_a_cup.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )

            if cup_non_trovati:
                st.warning(
                    "I seguenti CUP non risultano nel dataset ANAC (potrebbero riguardare "
                    "affidamenti non ordinari, essere di recente emissione o non ancora "
                    "propagati nel dataset open data): " + ", ".join(cup_non_trovati)
                )
else:
    st.info("Incolla uno o più CUP, oppure carica un file Excel, per iniziare.")

st.markdown("---")
st.caption(
    "Dati recuperati dai dataset open data e dall'API pubblica dell'Autorità Nazionale "
    "Anticorruzione (ANAC). L'applicazione non memorizza né trasmette i dati a terzi: "
    "l'elaborazione avviene solo durante la sessione attiva."
)
